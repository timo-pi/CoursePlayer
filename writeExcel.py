from openpyxl import Workbook, load_workbook
import os

def createReport(path):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50

    # Rows can also be appended
    ws.append(["COURSE FILE PATH", "SCORM VERSION", "ONE ITEM CHECK", "ADLNAV NAMESPACE (2004 4th only)", "SPECIAL CHARACTERS CHECK", "OVERALL STATUS"])
    try:
        wb.save(os.path.join(path, "scorm_tester_report.xlsx"))
        wb.close()
    except:
        print("Not possible to create SCORM-Report - file maybe open?")

def writeReport(path, data):
    try:
        wb = load_workbook(os.path.join(path, "scorm_tester_report.xlsx"))
        ws = wb.active
        ws.append(data)
        wb.save(os.path.join(path, "scorm_tester_report.xlsx"))
        wb.close()
    except:
        print("Not possible to save Test-Report - file maybe open?")

def createItemsReport(path, data):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 120
    ws.append(['SPECIAL CHARACTERS PRESENT IN IMSMANIFEST.XML'])
    for i in data:
        row = []
        row.append(i)
        ws.append(row)
    try:
        wb.save(path)
        wb.close()
    except:
        print("Not possible to write items-report - file maybe open?")

